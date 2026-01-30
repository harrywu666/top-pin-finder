"""
Excel导出模块
负责将下载的图片信息导出到Excel表格
"""

import os
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
from io import BytesIO
from ..core.logger import logger


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self, output_path: str = "./downloads/download_report.xlsx"):
        """
        初始化Excel导出器
        
        Args:
            output_path: 输出Excel文件路径
        """
        self.output_path = output_path
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Pinterest下载记录"
        
        # 初始化表头
        self._setup_headers()
        
        # 当前行号
        self.current_row = 2
        
        logger.info(f"Excel导出器已初始化，输出路径: {output_path}")
    
    def _setup_headers(self):
        """设置表头"""
        headers = ["图片预览", "点赞数", "原图链接", "本地文件名"]
        
        # 设置表头
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
        
        # 设置列宽
        self.worksheet.column_dimensions['A'].width = 30  # 图片列
        self.worksheet.column_dimensions['B'].width = 15  # 点赞数列
        self.worksheet.column_dimensions['C'].width = 50  # 链接列
        self.worksheet.column_dimensions['D'].width = 30  # 文件名列
    
    def add_record(self, image_path: str, likes: int, original_url: str, filename: str, auto_save: bool = False):
        """
        添加一条下载记录
        
        Args:
            image_path: 图片本地路径
            likes: 点赞数
            original_url: 原图链接
            filename: 本地文件名
            auto_save: 是否添加后立即保存(默认False,可在下载完成时设为True以实时保存)
        """
        try:
            row = self.current_row
            
            # 1. 插入缩略图
            if os.path.exists(image_path):
                # 创建缩略图
                img = Image.open(image_path)
                
                # 调整大小以适应单元格 (最大宽度200px)
                max_width = 200
                max_height = 150
                img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                
                # 保存到临时BytesIO
                img_byte_arr = BytesIO()
                img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                
                # 插入到Excel
                xl_img = XLImage(img_byte_arr)
                
                # 设置图片位置
                cell_ref = f'A{row}'
                xl_img.anchor = cell_ref
                
                self.worksheet.add_image(xl_img)
                
                # 设置行高以容纳图片 (单位是点，1点约等于1.33像素)
                self.worksheet.row_dimensions[row].height = max_height * 0.75
            
            # 2. 填写点赞数
            likes_cell = self.worksheet.cell(row=row, column=2)
            likes_cell.value = likes
            likes_cell.alignment = Alignment(horizontal='center', vertical='center')
            likes_cell.font = Font(size=11)
            
            # 3. 填写原图链接（设置为超链接）
            url_cell = self.worksheet.cell(row=row, column=3)
            url_cell.value = original_url
            url_cell.hyperlink = original_url
            url_cell.font = Font(color="0563C1", underline="single", size=10)
            url_cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # 4. 填写本地文件名
            filename_cell = self.worksheet.cell(row=row, column=4)
            filename_cell.value = filename
            filename_cell.alignment = Alignment(vertical='center')
            filename_cell.font = Font(size=10)
            
            self.current_row += 1
            logger.debug(f"已添加Excel记录: {filename}")
            
            # 如果启用自动保存,立即保存Excel文件
            if auto_save:
                self.save()
                logger.debug(f"Excel已自动保存(当前 {self.get_record_count()} 条记录)")
            
        except Exception as e:
            logger.error(f"添加Excel记录失败: {e}")
    
    def save(self):
        """保存Excel文件"""
        try:
            # 确保输出目录存在
            output_dir = os.path.dirname(self.output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 保存工作簿
            self.workbook.save(self.output_path)
            logger.info(f"✓ Excel报告已保存: {self.output_path}")
            
            # 重新加载workbook以支持多次保存
            # 这样可以避免"I/O operation on closed file"错误
            self.workbook = load_workbook(self.output_path)
            self.worksheet = self.workbook.active
            
            return True
            
        except Exception as e:
            logger.error(f"保存Excel文件失败: {e}")
            return False
    
    def get_record_count(self) -> int:
        """获取记录数量"""
        return self.current_row - 2  # 减去表头行


if __name__ == "__main__":
    # 测试Excel导出器
    exporter = ExcelExporter("./test_report.xlsx")
    
    # 添加测试记录
    exporter.add_record(
        image_path="./downloads/landscape_3296_0001.jpg",
        likes=3296,
        original_url="https://i.pinimg.com/originals/test.jpg",
        filename="landscape_3296_0001.jpg"
    )
    
    # 保存
    exporter.save()
    print(f"测试完成，已创建 {exporter.get_record_count()} 条记录")
